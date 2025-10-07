"""
Excel Processor
===============

Processes multi-sheet Excel files (XLSX/XLS) into hierarchical structure:
- Parent: Excel file asset
- Children: Sheet assets (CSV kind)
- Grandchildren: Row assets (CSV_ROW kind)
"""

import asyncio
import csv
import logging
from typing import List, Dict, Any, Tuple, Optional
from app.models import Asset, AssetKind
from app.schemas import AssetCreate
from .base import BaseProcessor, ProcessingError

logger = logging.getLogger(__name__)


class ExcelProcessor(BaseProcessor):
    """
    Process multi-sheet Excel files.
    
    Creates a hierarchical structure:
    Excel → Sheets → Rows
    """
    
    def can_process(self, asset: Asset) -> bool:
        """Check if asset is a processable Excel file."""
        return (
            asset.kind == AssetKind.CSV and
            asset.blob_path is not None and
            asset.blob_path.endswith(('.xlsx', '.xls'))
        )
    
    async def process(self, asset: Asset) -> List[Asset]:
        """
        Process Excel file and create sheet and row assets.
        
        Args:
            asset: Parent Excel asset
            
        Returns:
            List of Sheet assets (which have their own Row children)
        """
        if not self.can_process(asset):
            raise ProcessingError(f"Cannot process asset {asset.id} as Excel")
        
        skip_rows = self.context.options.get('skip_rows', 0)
        max_rows = self.context.max_rows
        
        # Read file
        file_stream = await self.context.storage_provider.get_file(asset.blob_path)
        file_bytes = await asyncio.to_thread(file_stream.read)
        
        # Parse Excel file
        file_ext = asset.blob_path.split('.')[-1].lower()
        sheets_data = await self._parse_excel(file_bytes, file_ext, skip_rows)
        
        if not sheets_data:
            raise ProcessingError("Excel file contains no data")
        
        # Update parent asset summary
        asset.text_content = f"Excel workbook with {len(sheets_data)} sheet(s)"
        asset.source_metadata.update({
            'sheet_count': len(sheets_data),
            'sheet_names': [sheet['name'] for sheet in sheets_data],
            'total_rows': sum(sheet['row_count'] for sheet in sheets_data),
            'is_multisheet_excel': True,
            'processing_options': self.context.options
        })
        
        # Create sheet assets
        sheet_assets = []
        for sheet_index, sheet_data in enumerate(sheets_data):
            sheet_asset = await self._process_sheet(
                asset, sheet_data, sheet_index, max_rows
            )
            sheet_assets.append(sheet_asset)
        
        logger.info(
            f"Processed Excel file: {len(sheets_data)} sheets, "
            f"total {sum(sheet['row_count'] for sheet in sheets_data)} rows"
        )
        
        return sheet_assets
    
    async def _parse_excel(
        self, 
        file_bytes: bytes, 
        file_ext: str, 
        skip_rows: int
    ) -> List[Dict[str, Any]]:
        """
        Parse Excel file into structured sheet data.
        
        Returns:
            List of dicts with keys: name, csv_text, row_count
        """
        try:
            import openpyxl
            import io
            
            def parse_sync():
                wb = openpyxl.load_workbook(
                    io.BytesIO(file_bytes), 
                    read_only=True, 
                    data_only=True
                )
                
                sheets_data = []
                
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    
                    # Convert sheet to CSV format
                    csv_lines = []
                    row_count = 0
                    
                    for i, row in enumerate(sheet.iter_rows(values_only=True)):
                        if i < skip_rows:
                            continue
                        
                        # Convert values to strings
                        row_values = [str(cell) if cell is not None else '' for cell in row]
                        
                        # Skip empty rows
                        if any(val.strip() for val in row_values):
                            # Properly escape CSV values
                            escaped_values = []
                            for val in row_values:
                                # If value contains comma, newline, or quote, wrap in quotes and escape internal quotes
                                if ',' in val or '\n' in val or '"' in val:
                                    # Escape internal quotes by doubling them
                                    escaped_val = val.replace('"', '""')
                                    escaped_values.append(f'"{escaped_val}"')
                                else:
                                    escaped_values.append(val)
                            csv_lines.append(','.join(escaped_values))
                            row_count += 1
                    
                    if csv_lines:  # Only include non-empty sheets
                        sheets_data.append({
                            'name': sheet_name,
                            'csv_text': '\n'.join(csv_lines),
                            'row_count': row_count
                        })
                
                wb.close()
                return sheets_data
            
            sheets_data = await asyncio.to_thread(parse_sync)
            logger.info(f"Converted Excel file: {len(sheets_data)} sheets")
            return sheets_data
            
        except ImportError:
            raise ProcessingError(
                "openpyxl library not installed. Install with: pip install openpyxl"
            )
        except Exception as e:
            raise ProcessingError(f"Failed to parse Excel file: {e}")
    
    async def _process_sheet(
        self,
        parent_asset: Asset,
        sheet_data: Dict[str, Any],
        sheet_index: int,
        max_rows: int
    ) -> Asset:
        """
        Process a single sheet and create row assets.
        
        Args:
            parent_asset: Parent Excel asset
            sheet_data: Dict with 'name', 'csv_text', 'row_count'
            sheet_index: Index of sheet in workbook
            max_rows: Maximum rows to process
            
        Returns:
            Sheet asset with row children
        """
        sheet_name = sheet_data['name']
        csv_text = sheet_data['csv_text']
        
        # Create sheet asset
        sheet_asset_create = AssetCreate(
            title=sheet_name,
            kind=AssetKind.CSV,
            user_id=parent_asset.user_id,
            infospace_id=parent_asset.infospace_id,
            parent_asset_id=parent_asset.id,
            part_index=sheet_index,
            text_content="",  # Will be filled with row summaries
            source_metadata={
                'sheet_name': sheet_name,
                'sheet_index': sheet_index,
                'parent_excel_file': parent_asset.title,
                'row_count': sheet_data['row_count'],
                'is_excel_sheet': True
            }
        )
        
        # Parse CSV text
        csv_lines = csv_text.split('\n')
        all_rows = list(csv.reader(csv_lines, delimiter=','))
        
        if not all_rows:
            logger.warning(f"Sheet '{sheet_name}' has no rows")
            return self.context.asset_service.create_asset(sheet_asset_create)
        
        # Smart header detection: find the row with the most non-empty cells
        header_row_idx, header = self._detect_header_row(all_rows, sheet_name)
        
        if header_row_idx is None or not header:
            logger.warning(f"Sheet '{sheet_name}' has no valid header row")
            return self.context.asset_service.create_asset(sheet_asset_create)
        
        # Update metadata with detected header position
        sheet_asset_create.source_metadata['header_row_index'] = header_row_idx
        sheet_asset_create.source_metadata['data_starts_at_row'] = header_row_idx + 1
        
        # Process rows (skip rows before and including header)
        child_assets = []
        full_text_parts = [f"Sheet: {sheet_name}", f"Headers: {' | '.join(header)}"]
        rows_processed = 0
        
        # Start processing from the row after the header
        for idx, row in enumerate(all_rows[header_row_idx + 1:]):
            if rows_processed >= max_rows:
                logger.warning(f"Sheet '{sheet_name}' processing stopped at {max_rows} rows limit")
                break
            
            if not any(cell.strip() for cell in row if cell):
                continue
            
            # Normalize row length
            while len(row) < len(header):
                row.append('')
            if len(row) > len(header):
                row = row[:len(header)]
            
            # Clean row data
            cleaned_row = [cell.replace('\x00', '').strip() for cell in row]
            row_data = {header[j]: cleaned_row[j] for j in range(len(header))}
            row_text = ' | '.join(cleaned_row)
            
            # Debug: Log if row_text is empty or very short
            if not row_text or len(row_text.strip('| ')) < 3:
                logger.warning(f"Sheet '{sheet_name}' Row {rows_processed}: text_content is suspiciously empty: '{row_text}'")
            
            full_text_parts.append(row_text)
            
            # Generate title: {sheet_name} | {index} | {first_cols}
            title_parts = [sheet_name, str(rows_processed + 1)]
            title_parts.extend(
                v[:25] + ('...' if len(v) > 25 else '')
                for v in cleaned_row[:2] if v.strip()
            )
            row_title = (
                " | ".join(title_parts) if len(title_parts) > 2
                else f"{sheet_name} Row {rows_processed + 1}"
            )
            
            # Store row data for later creation (after sheet asset is created)
            row_metadata = {
                'title': row_title,
                'text_content': row_text,
                'part_index': rows_processed,
                'original_row_data': row_data
            }
            child_assets.append(row_metadata)
            rows_processed += 1
            
            # Yield periodically
            if rows_processed % 1000 == 0:
                await asyncio.sleep(0.01)
        
        # Update sheet asset with content summary
        sheet_asset_create.text_content = "\n".join(full_text_parts)
        sheet_asset_create.source_metadata.update({
            'columns': header,
            'column_count': len(header),
            'rows_processed': rows_processed
        })
        
        # Create the sheet asset first
        sheet_asset = self.context.asset_service.create_asset(sheet_asset_create)
        
        # Now create row assets as children of the sheet
        for row_metadata in child_assets:
            child_asset_create = AssetCreate(
                title=row_metadata['title'],
                kind=AssetKind.CSV_ROW,
                user_id=parent_asset.user_id,
                infospace_id=parent_asset.infospace_id,
                parent_asset_id=sheet_asset.id,  # Set sheet as parent
                part_index=row_metadata['part_index'],
                text_content=row_metadata['text_content'],
                source_metadata={
                    'sheet_name': sheet_name,
                    'sheet_index': sheet_index,
                    'row_number': row_metadata['part_index'] + 1,
                    'data_row_index': row_metadata['part_index'],
                    'original_row_data': row_metadata['original_row_data'],
                    'excel_file': parent_asset.title
                }
            )
            self.context.asset_service.create_asset(child_asset_create)
        
        logger.info(f"Processed sheet '{sheet_name}': {rows_processed} rows, {len(header)} columns, created {len(child_assets)} row assets")
        
        return sheet_asset
    
    def _detect_header_row(self, all_rows: List[List[str]], sheet_name: str) -> Tuple[Optional[int], List[str]]:
        """
        Intelligently detect which row is the header in an Excel sheet.
        
        Strategy:
        1. Skip empty rows at the start
        2. Find the row with the most non-empty cells (likely header)
        3. Skip rows that look like titles (1-2 non-empty cells)
        4. Verify the next row has similar number of cells (data row validation)
        
        Args:
            all_rows: All rows from the CSV conversion
            sheet_name: Sheet name for logging
            
        Returns:
            Tuple of (header_row_index, header_list) or (None, []) if not found
        """
        if not all_rows:
            return None, []
        
        # Scan first 20 rows to find the header
        scan_limit = min(20, len(all_rows))
        
        # Count non-empty cells in each row
        row_scores = []
        for idx in range(scan_limit):
            row = all_rows[idx]
            non_empty_count = sum(1 for cell in row if cell.strip())
            
            # Skip completely empty rows
            if non_empty_count == 0:
                continue
            
            # Skip rows with very few cells (likely titles/labels)
            if non_empty_count <= 2:
                continue
            
            # Calculate "header-likeness" score
            # - More non-empty cells = better
            # - Text length variance (headers are usually similar length)
            cell_lengths = [len(cell.strip()) for cell in row if cell.strip()]
            avg_length = sum(cell_lengths) / len(cell_lengths) if cell_lengths else 0
            
            # Headers typically have 5-30 character column names
            length_score = 1.0 if 5 <= avg_length <= 30 else 0.5
            
            score = non_empty_count * length_score
            row_scores.append((idx, non_empty_count, score, row))
        
        if not row_scores:
            logger.warning(f"Sheet '{sheet_name}': No potential header rows found")
            return None, []
        
        # Sort by score (descending) and take the best candidate
        row_scores.sort(key=lambda x: x[2], reverse=True)
        header_idx, cell_count, score, raw_header_row = row_scores[0]
        
        # Validate: Check if the next few rows have similar cell counts (data rows)
        if header_idx + 1 < len(all_rows):
            next_row_cells = sum(1 for cell in all_rows[header_idx + 1] if cell.strip())
            
            # Data rows should have similar number of cells as header
            # Allow some variance (e.g., 50% of header columns)
            if next_row_cells < cell_count * 0.5:
                logger.warning(
                    f"Sheet '{sheet_name}': Detected header at row {header_idx + 1} "
                    f"but next row has too few cells ({next_row_cells} vs {cell_count})"
                )
                # Try second best candidate if available
                if len(row_scores) > 1:
                    header_idx, cell_count, score, raw_header_row = row_scores[1]
        
        # Process header: keep all columns, name empty ones
        header = [
            h.strip() if h.strip() else f'Column_{i+1}' 
            for i, h in enumerate(raw_header_row)
        ]
        
        logger.info(
            f"Sheet '{sheet_name}': Detected header at row {header_idx + 1} "
            f"with {cell_count} non-empty columns"
        )
        
        return header_idx, header

